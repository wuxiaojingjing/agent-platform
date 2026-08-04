#!/usr/bin/env python3
"""从素材解析下的 Excel 重生成 tech-domains.yaml / menu-tree.yaml。"""

from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[2]
AGENT = ROOT / "mobile-banking-assistant-agent"

TECH_CODE = {
    "理财服务": "wealth_product",
    "基金服务": "fund_service",
    "贵金属服务": "precious_metal",
    "债券服务": "bond_service",
    "外汇服务": "fx_service",
    "存款服务": "deposit_service",
    "保险服务": "insurance_service",
    "财富聚合服务": "wealth_aggregate",
    "贷款服务": "loan_service",
    "安全服务": "security_service",
    "个人信息管理": "personal_info",
    "金融助手": "finance_assistant",
    "渠道设置": "channel_settings",
    "咨询服务": "advisory_service",
    "客户专享": "vip_service",
    "薪资服务": "payroll_service",
    "民生服务": "livelihood_service",
    "企业服务": "enterprise_service",
    "权益运营": "benefits_ops",
    "生活服务": "life_service",
    "数字人民币": "e_cny",
    "网点服务": "branch_service",
    "信用卡服务": "creditcard_service",
    "账户管理": "account",
    "支付管理": "payment",
    "转账服务": "transfer",
}

# payment 已是「支付管理」规范码；转账写 transfer，勿把 payment 别名到 transfer
ALIASES = {
    "wealth": "wealth_aggregate",
    "fund": "fund_service",
    "insurance": "insurance_service",
    "creditcard": "creditcard_service",
}

LABEL_MAP = {
    "理财": "wealth_product",
    "基金": "fund_service",
    "贵金属": "precious_metal",
    "债券": "bond_service",
    "外汇": "fx_service",
    "存款": "deposit_service",
    "保险": "insurance_service",
    "财富": "wealth_aggregate",
    "资产": "wealth_aggregate",
    "贷款": "loan_service",
    "安全中心": "security_service",
    "安全": "security_service",
    "个人信息": "personal_info",
    "设置": "channel_settings",
    "渠道": "channel_settings",
    "咨询": "advisory_service",
    "客服": "advisory_service",
    "客户专享": "vip_service",
    "薪管家": "payroll_service",
    "薪资": "payroll_service",
    "民生": "livelihood_service",
    "社保": "livelihood_service",
    "医保": "livelihood_service",
    "公积金": "livelihood_service",
    "企业": "enterprise_service",
    "小微": "enterprise_service",
    "权益": "benefits_ops",
    "i豆": "benefits_ops",
    "生活": "life_service",
    "缴费": "life_service",
    "数字人民币": "e_cny",
    "e钱包": "e_cny",
    "网点": "branch_service",
    "云网点": "branch_service",
    "信用卡": "creditcard_service",
    "账户": "account",
    "收支": "account",
    "支付": "payment",
    "e支付": "payment",
    "转账": "transfer",
    "转账汇款": "transfer",
    "汇款": "transfer",
    "收款": "transfer",
}

BIZ_DEFAULT = {
    "财富管理": "wealth_aggregate",
    "贷款服务": "loan_service",
    "金融助手": "finance_assistant",
    "客户专享": "vip_service",
    "民生服务": "livelihood_service",
    "企业服务": "enterprise_service",
    "权益服务": "benefits_ops",
    "生活服务": "life_service",
    "数字人民币": "e_cny",
    "网点服务": "branch_service",
    "信用卡": "creditcard_service",
    "账户管理": "account",
    "支付管理": "payment",
    "转账服务": "transfer",
}


class NoAliasDumper(yaml.SafeDumper):
    def ignore_aliases(self, data):
        return True


def infer_tech(biz, l1, l2, l3, final):
    for p in (l3, l2, l1, final):
        if not p:
            continue
        text = str(p).strip()
        for label, code in LABEL_MAP.items():
            if label in text:
                return code
    return BIZ_DEFAULT.get(biz, "unmapped")


def slugify(name: str) -> str:
    return re.sub(r"\s+", "", name)


def dump(doc, path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        f.write("# 自动从 Excel 生成，勿手改条目；改源表后重跑 scripts/import_domain_menu_assets.py\n")
        yaml.dump(doc, f, allow_unicode=True, sort_keys=False, Dumper=NoAliasDumper, width=120)


def main():
    wb = openpyxl.load_workbook(ROOT / "素材解析/业务领域映射科技领域.xlsx", data_only=True)
    ws = wb.active
    domains = []
    cur_biz = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        b, t = row[0], row[1]
        if b and str(b).strip():
            cur_biz = str(b).strip()
        if not t:
            continue
        name = str(t).strip()
        domains.append({
            "code": TECH_CODE[name],
            "name": name,
            "businessDomain": cur_biz,
        })

    dump({
        "version": "1.0.0",
        "source": "素材解析/业务领域映射科技领域.xlsx",
        "aliases": ALIASES,
        "domains": domains,
    }, AGENT / "assets/domains/tech-domains.yaml")

    wb2 = openpyxl.load_workbook(ROOT / "素材解析/爱存不存_菜单树.xlsx", data_only=True)
    ws2 = wb2.active
    menus = []
    seen = set()
    for row in ws2.iter_rows(min_row=2, values_only=True):
        vals = list(row) + [None] * 8
        b, a1, a2, a3, a4, bks, final, typ = vals[:8]
        if final is None or str(final).strip() == "":
            continue
        biz = str(b).strip() if b else ""
        final_name = str(final).strip()
        typ_s = str(typ).strip() if typ else "菜单"
        if typ_s == "芝单":
            typ_s = "菜单"
        kind = "workflow" if "工作流" in typ_s else "menu"
        l1 = str(a1).strip() if a1 else ""
        l2 = str(a2).strip() if a2 else ""
        l3 = str(a3).strip() if a3 else ""
        l4 = str(a4).strip() if a4 else ""
        bks_s = str(bks).strip() if bks else ""
        tech = infer_tech(biz, l1, l2, l3, final_name)
        base = f"menu.{tech}.{slugify(final_name)}"
        menu_id = base
        n = 2
        while menu_id in seen:
            menu_id = f"{base}_{n}"
            n += 1
        seen.add(menu_id)
        path = "＞".join(x for x in (l1, l2, l3, l4) if x)
        menus.append({
            "menuId": menu_id,
            "businessDomain": biz,
            "techDomain": tech,
            "path": path,
            "finalName": final_name,
            "bksPath": bks_s,
            "kind": kind,
        })

    dump({
        "version": "1.0.0",
        "source": "素材解析/爱存不存_菜单树.xlsx",
        "menus": menus,
    }, AGENT / "assets/menus/menu-tree.yaml")

    print(f"domains={len(domains)} menus={len(menus)} {Counter(m['kind'] for m in menus)}")


if __name__ == "__main__":
    main()
